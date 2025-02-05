import os
import openpyxl

os.chdir('modulo_9' + os.sep + 'aula10')
planilha = openpyxl.load_workbook('pessoas.xlsx')
print(planilha.sheetnames)
sheet1 = planilha.get_sheet_by_name('Sheet1')
print(sheet1['B4'].value)

sheet1['B4'].value = 'Jimmy'

for linha in sheet1.iter_rows(min_row=2, max_row=10, min_col=2):
    print(linha[0].value,linha[1].value,linha[2].value, linha[3].value, linha[4].value, linha[5].value, linha[6].value)

for linha in sheet1.iter_cols(min_col=3, max_col=3, min_row=2):
    for cell in linha:
        print(cell.value)